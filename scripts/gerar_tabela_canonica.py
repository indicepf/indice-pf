# -*- coding: utf-8 -*-
"""
Gera tabela_canonica_ingredientes.xlsx a partir de:
  - Tabela_Indice_PF_Ataulizada.xlsx  (101 pratos, 925 linhas)
  - Ingredientes_Scrapping.xlsx        (mapa variação -> nome canônico bruto)
  - scripts/mapa_canonico.py           (decisões manuais: ATOMICO / COMPOSTO)

Valida que todos os nomes canônicos brutos estão cobertos antes de exportar.
Execução única/manual — não faz parte do pipeline semanal.
"""
import os, re, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, os.path.dirname(__file__))
from mapa_canonico import BASE, ATOMICO, COMPOSTO, REVISAR, consolidar, PRATO_ALIAS
from tripe_scraping import TRIPE

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TAB  = os.path.join(RAIZ, "data", "Tabela_Indice_PF_Ataulizada.xlsx")
SCR  = os.path.join(RAIZ, "data", "Ingredientes_Scrapping.xlsx")
OUT  = os.path.join(RAIZ, "data", "tabela_canonica_ingredientes.xlsx")


def parse_qtd_g(txt):
    """Extrai gramas/ml de strings como '80g', '1 un (55g)', '120g (...)+10g (...)'."""
    if not txt:
        return None
    total = 0.0
    achou = False
    for m in re.finditer(r'(\d+[.,]?\d*)\s*(g|ml)\b', txt.lower()):
        total += float(m.group(1).replace(',', '.'))
        achou = True
    if achou:
        return round(total, 2)
    # fallback: "1 un" sem peso -> None (revisar manualmente)
    return None


def carregar_var2canon():
    wb = openpyxl.load_workbook(SCR, data_only=True)
    ws = wb["Lista Scrapping"]
    m = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        canon = row[1].strip()
        for v in row[2].split("|"):
            m[v.strip()] = canon
    return m


def carregar_linhas(var2canon):
    wb = openpyxl.load_workbook(TAB, data_only=True)
    linhas = []
    for regiao in wb.sheetnames:
        ws = wb[regiao]
        for r in ws.iter_rows(min_row=2, values_only=True):
            prato, ing, qbruta, aprov = r
            if not ing:
                continue
            ing = ing.strip()
            prato_nome = prato.strip() if prato else ""
            prato_nome = PRATO_ALIAS.get(prato_nome, prato_nome)
            pb = parse_qtd_g(qbruta)
            pc = parse_qtd_g(aprov)
            meta, compra = meta_e_compra(qbruta, pb, pc)
            linhas.append({
                "regiao": regiao,
                "prato": prato_nome,
                "ingrediente_raw": ing,
                "canon": var2canon.get(ing),
                "qtd_txt": qbruta.strip() if qbruta else "",
                "qtd_pb_g": pb,           # receita crua original (PB)
                "qtd_cozida_g": pc,       # rendimento do PB (PC)
                "qtd_meta_g": meta,       # meta servida no prato
                "qtd_g": compra,          # COMPRA necessária — base do custo em todo o app
            })
    return linhas


def meta_e_compra(pb_txt, pb, pc):
    """Modelo por meta servida (planilha do sócio, 12/07/2026): FC = PB/PC.

    - Itens que ENCOLHEM ao cozinhar (FC > 1, carnes): a quantidade da receita
      era a porção desejada NO PRATO → meta = PB, compra = PB × FC.
    - Itens que EXPANDEM (FC <= 1, arroz/feijão) e itens por UNIDADE (ovos):
      a quantidade da receita já era a compra → compra = PB, meta = PC.
    - Sem PC parseável: compra = PB, meta indefinida.
    """
    if pb is None:
        return None, None
    if pc is None or pc <= 0:
        return None, pb
    por_unidade = bool(re.search(r'\bun\b|\bunid', (pb_txt or '').lower()))
    fc = pb / pc
    if fc > 1.001 and not por_unidade:
        return pb, round(pb * fc, 2)
    return pc, pb


def validar(canons):
    """Garante que todo canônico bruto está em ATOMICO ou COMPOSTO, e que
    toda base referenciada existe em BASE e proporções somam ~1."""
    faltando, prop_ruim, base_inexistente = [], [], set()
    for c in canons:
        if c not in ATOMICO and c not in COMPOSTO:
            faltando.append(c)
        if c in ATOMICO and ATOMICO[c] not in BASE:
            base_inexistente.add(ATOMICO[c])
    for c, comps in COMPOSTO.items():
        s = sum(p for _, p in comps)
        if abs(s - 1.0) > 0.001:
            prop_ruim.append((c, round(s, 3)))
        for b, _ in comps:
            if b not in BASE:
                base_inexistente.add(b)
    return faltando, prop_ruim, base_inexistente


def base_de(canon):
    """Retorna lista de (base consolidada, proporção) para um canônico bruto."""
    if canon in ATOMICO:
        return [(consolidar(ATOMICO[canon]), 1.0)]
    return [(consolidar(b), p) for b, p in COMPOSTO[canon]]


def estilizar_header(ws):
    fill = PatternFill("solid", fgColor="1A1A1A")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")


def main():
    var2canon = carregar_var2canon()
    linhas = carregar_linhas(var2canon)
    canons = sorted({l["canon"] for l in linhas if l["canon"]})

    faltando, prop_ruim, base_inexistente = validar(canons)
    if faltando or prop_ruim or base_inexistente:
        print("VALIDAÇÃO FALHOU — corrija mapa_canonico.py antes de exportar:")
        if faltando:
            print(f"\n  {len(faltando)} canônicos sem mapeamento:")
            for c in faltando:
                print("    -", c)
        if base_inexistente:
            print(f"\n  bases referenciadas que não existem em BASE:")
            for b in sorted(base_inexistente):
                print("    -", b)
        if prop_ruim:
            print(f"\n  compostos com proporção != 1.0:")
            for c, s in prop_ruim:
                print(f"    - {c}: soma={s}")
        sys.exit(1)

    print(f"OK: {len(canons)} canônicos cobertos, proporções válidas.")

    wb = openpyxl.Workbook()

    # ── Aba 1: Ingredientes Canônicos ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Ingredientes Canônicos"
    ws1.append(["Ingrediente canônico (base)", "Categoria", "Nº de pratos", "Nº de ocorrências (linhas)"])
    # conta uso de cada base nas receitas decompostas
    uso_linhas, uso_pratos = {}, {}
    for l in linhas:
        if not l["canon"]:
            continue
        for b, _ in base_de(l["canon"]):
            uso_linhas[b] = uso_linhas.get(b, 0) + 1
            uso_pratos.setdefault(b, set()).add((l["regiao"], l["prato"]))
    for base in sorted(uso_linhas):
        if base in ("Água/Subproduto (sem custo)", "SEM COTAÇÃO (revisar)"):
            continue
        ws1.append([base, BASE.get(base, "?"), len(uso_pratos.get(base, set())), uso_linhas.get(base, 0)])
    estilizar_header(ws1)

    # ── Aba 2: Receitas Decompostas ──────────────────────────────────────────
    ws2 = wb.create_sheet("Receitas Decompostas")
    ws2.append(["Região", "Prato", "Ingrediente original", "Qtd original",
                "Compra (g/ml)", "Ingrediente canônico (base)", "Compra atribuída (g/ml)", "Origem"])
    for l in linhas:
        comps = base_de(l["canon"]) if l["canon"] else [("(SEM CANÔNICO)", 1.0)]
        origem = "direto" if l["canon"] in ATOMICO else "decomposto"
        for b, prop in comps:
            qtd_atrib = round(l["qtd_g"] * prop, 2) if l["qtd_g"] is not None else None
            ws2.append([l["regiao"], l["prato"], l["ingrediente_raw"], l["qtd_txt"],
                        l["qtd_g"], b, qtd_atrib, origem])
    estilizar_header(ws2)

    # ── Aba 3: Decomposições Aplicadas ───────────────────────────────────────
    ws3 = wb.create_sheet("Decomposições Aplicadas")
    ws3.append(["Composto (nome bruto)", "Ingrediente-base", "Proporção", "Observação"])
    for c in sorted(COMPOSTO):
        for i, (b, p) in enumerate(COMPOSTO[c]):
            obs = REVISAR.get(c, "") if i == 0 else ""
            ws3.append([c if i == 0 else "", b, p, obs])
    estilizar_header(ws3)

    # ── Aba 4: Fusões Aplicadas ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Fusões Aplicadas")
    ws4.append(["Nome canônico bruto (404)", "→ Base final / decomposição", "Tipo", "Nº ocorrências"])
    ocor = {}
    for l in linhas:
        ocor[l["canon"]] = ocor.get(l["canon"], 0) + 1
    for c in canons:
        if c in ATOMICO:
            destino, tipo = ATOMICO[c], "atômico"
        else:
            destino = " + ".join(f"{b} ({int(p*100)}%)" for b, p in COMPOSTO[c])
            tipo = "composto"
        ws4.append([c, destino, tipo, ocor.get(c, 0)])
    estilizar_header(ws4)

    # ── Aba 5: A Revisar ─────────────────────────────────────────────────────
    ws5 = wb.create_sheet("A Revisar")
    ws5.append(["Nome canônico bruto", "Destino atual", "Motivo / o que revisar"])
    for c in sorted(REVISAR):
        if c in ATOMICO:
            destino = ATOMICO[c]
        elif c in COMPOSTO:
            destino = " + ".join(f"{b} ({int(p*100)}%)" for b, p in COMPOSTO[c])
        else:
            destino = "?"
        ws5.append([c, destino, REVISAR[c]])
    estilizar_header(ws5)

    # ── Aba 0a: Receitas Finais (prato × ingrediente canônico × porção) ──────
    # Consolida: soma a quantidade quando o mesmo ingrediente-base aparece em
    # mais de uma linha do prato. Remove água/subproduto (não é item de compra).
    from collections import defaultdict
    CAMPOS = ("qtd_g", "qtd_pb_g", "qtd_cozida_g", "qtd_meta_g")
    consol = {c: defaultdict(float) for c in CAMPOS}
    for l in linhas:
        if not l["canon"]:
            continue
        for b, prop in base_de(l["canon"]):
            if b == "Água/Subproduto (sem custo)":
                continue
            for c in CAMPOS:
                consol[c][(l["regiao"], l["prato"], b)] += (l[c] or 0) * prop

    ws_f = wb.create_sheet("Receitas Finais", 0)
    ws_f.append(["Região", "Prato", "Ingrediente canônico", "Categoria",
                 "Compra/porção (g ou ml)", "PB receita (g)", "PC rendimento (g)", "Meta no prato (g)", "Obs"])
    for (regiao, prato, base), qtd in sorted(consol["qtd_g"].items()):
        tb = TRIPE.get(base, {})
        if tb.get("unidade") == "fixo":
            obs = f"preço fixo R${tb.get('custo_fixo', 0):.2f}"
        elif base == "SEM COTAÇÃO (revisar)":
            obs = "SEM COTAÇÃO — revisar"
        else:
            obs = ""
        k = (regiao, prato, base)
        ws_f.append([regiao, prato, base, BASE.get(base, "?"), round(qtd, 1),
                     round(consol["qtd_pb_g"][k], 1), round(consol["qtd_cozida_g"][k], 1),
                     round(consol["qtd_meta_g"][k], 1), obs])
    estilizar_header(ws_f)

    # ── Aba 0b: Pratos (resumo) ──────────────────────────────────────────────
    por_prato = defaultdict(lambda: {"ing": 0, "compra": 0.0, "meta": 0.0})
    for (regiao, prato, base), qtd in consol["qtd_g"].items():
        por_prato[(regiao, prato)]["ing"] += 1
        por_prato[(regiao, prato)]["compra"] += qtd
        por_prato[(regiao, prato)]["meta"] += consol["qtd_meta_g"][(regiao, prato, base)]
    ws_p = wb.create_sheet("Pratos (resumo)", 1)
    ws_p.append(["Região", "Prato", "Nº ingredientes canônicos", "Compra/porção (g)", "Meta no prato/porção (g)"])
    for (regiao, prato), d in sorted(por_prato.items()):
        ws_p.append([regiao, prato, d["ing"], round(d["compra"], 1), round(d["meta"], 1)])
    estilizar_header(ws_p)

    # ── Aba: Scraping (proposta de tripé) ────────────────────────────────────
    ws_s = wb.create_sheet("Scraping (99 termos)", 2)
    ws_s.append(["Ingrediente canônico", "Categoria", "Termo de busca", "Unidade",
                 "Peso ref. (g) p/ maço", "Custo fixo (R$)", "Preço manual (R$/kg)",
                 "palavras_ok (≥1)", "palavras_nao (nenhuma)"])
    faltam_tripe = [b for b in sorted(uso_linhas)
                    if b not in ("Água/Subproduto (sem custo)", "SEM COTAÇÃO (revisar)")
                    and b not in TRIPE]
    for base in sorted(b for b in uso_linhas if b in TRIPE):
        t = TRIPE[base]
        ws_s.append([base, BASE.get(base, "?"), t["busca"], t["unidade"],
                     t.get("peso_ref_g", ""), t.get("custo_fixo", ""), t.get("preco_manual", ""),
                     " | ".join(t["ok"]), " | ".join(t["nao"])])
    estilizar_header(ws_s)
    if faltam_tripe:
        print("AVISO: bases sem tripé:", faltam_tripe)

    # larguras
    for ws in wb.worksheets:
        for col in ws.columns:
            largura = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(largura + 2, 60)

    wb.save(OUT)
    sentinelas = ("Água/Subproduto (sem custo)", "SEM COTAÇÃO (revisar)")
    bases_usadas = [b for b in uso_linhas if b not in sentinelas]
    fixos   = [b for b in bases_usadas if TRIPE.get(b, {}).get("unidade") == "fixo"]
    manuais = [b for b in bases_usadas if TRIPE.get(b, {}).get("preco_manual") is not None]
    scrapeaveis = [b for b in bases_usadas if b not in fixos and b not in manuais]
    print(f"Planilha gerada: {OUT}")
    print(f"  Ingredientes-base: {len(bases_usadas)} "
          f"({len(scrapeaveis)} scrapeáveis + {len(fixos)} preço fixo + {len(manuais)} preço manual)")
    print(f"  Orçamento: 250 chamadas/mês ÷ {len(scrapeaveis)} = {250 // len(scrapeaveis)} coletas completas/mês")
    print(f"  Linhas decompostas (aba 2): {ws2.max_row - 1}")
    print(f"  Compostos decompostos: {len(COMPOSTO)}")
    print(f"  Itens a revisar: {len(REVISAR)}")


if __name__ == "__main__":
    main()
